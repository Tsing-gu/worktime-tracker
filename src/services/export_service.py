# -*- coding: utf-8 -*-
"""
export_service - 数据导出服务
===============================

提供 CSV 和 Excel 两种格式的工时数据导出功能。
从 database 读取记录后生成文件，保存到桌面。

版本: 0.4.2
"""

import csv
import os
from datetime import date

from src.config import EXPORT_DIR
from src.data import database


# ─── CSV 导出 ────────────────────────────────────────────────

def export_csv(start_date: date, end_date: date, filepath: str = None) -> str:
    """
    导出指定日期范围的工时记录为 CSV 文件。

    使用 UTF-8-SIG 编码（带 BOM），确保 Excel 直接打开不乱码。

    Args:
        start_date: 起始日期
        end_date:   结束日期
        filepath:   输出路径（默认 ~/Desktop/工时记录_起_止.csv）

    Returns:
        文件保存路径
    """
    records = database.get_date_range_worktime(start_date, end_date)
    if not filepath:
        filepath = os.path.join(
            EXPORT_DIR,
            f"工时记录_{start_date.isoformat()}_to_{end_date.isoformat()}.csv",
        )

    # 表头
    headers = ["日期", "上班时间", "下班时间", "工时(小时)", "每日要求(小时)", "是否达标", "请假类型", "数据来源", "备注"]

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for r in records:
            start = r.get("start_time", "") or ""
            end = r.get("end_time", "") or ""
            total = r.get("total_hours") or 0
            # 优先用记录中持久化的 required_hours，NULL 时用 0
            req = r.get("required_hours") or 0
            reached = "是" if total >= req else "否"
            leave = r.get("leave_type") or ""
            if leave == "none":
                leave = ""
            source = r.get("source") or "auto"
            note = r.get("note") or ""
            writer.writerow([
                r["work_date"], start, end, f"{total:.2f}", f"{req:.1f}",
                reached, leave, source, note,
            ])

    return filepath


# ─── Excel 导出 ──────────────────────────────────────────────

def export_excel(start_date: date, end_date: date, filepath: str = None) -> str:
    """
    导出指定日期范围的工时记录为 Excel 文件（带格式）。

    格式特性:
        - 表头蓝底白字加粗
        - 达标行绿色、不足行红色、请假行蓝色
        - 居中对齐 + 细边框

    Args:
        start_date: 起始日期
        end_date:   结束日期
        filepath:   输出路径（默认 ~/Desktop/工时记录_起_止.xlsx）

    Returns:
        文件保存路径
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    records = database.get_date_range_worktime(start_date, end_date)
    if not filepath:
        filepath = os.path.join(
            EXPORT_DIR,
            f"工时记录_{start_date.isoformat()}_to_{end_date.isoformat()}.xlsx",
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "工时记录"

    # 表头
    headers = ["日期", "上班时间", "下班时间", "工时(小时)", "每日要求(小时)", "是否达标", "请假类型", "数据来源", "备注"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    # 写表头
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # 条件格式填充色
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    # 写数据行
    for row_idx, r in enumerate(records, 2):
        start = r.get("start_time", "") or ""
        end = r.get("end_time", "") or ""
        total = r.get("total_hours") or 0
        req = r.get("required_hours") or 0
        reached = total >= req
        leave = r.get("leave_type") or ""
        if leave == "none":
            leave = ""
        source = r.get("source") or "auto"
        note = r.get("note") or ""

        row_data = [
            r["work_date"], start, end, round(total, 2), round(req, 1),
            "是" if reached else "否", leave, source, note,
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = center_align
            cell.border = thin_border

        # 条件着色
        if leave:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = blue_fill
        elif reached:
            ws.cell(row=row_idx, column=6).fill = green_fill
        else:
            ws.cell(row=row_idx, column=6).fill = red_fill

    # 列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 18

    wb.save(filepath)
    return filepath
